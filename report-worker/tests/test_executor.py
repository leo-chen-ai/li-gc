import hashlib
import json
import sys
from contextlib import contextmanager, nullcontext
from datetime import date, timedelta
from pathlib import Path
from zipfile import ZIP_DEFLATED, ZipFile

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

import report_worker.executor as executor_module
import converter
import downloader
import target_login
import uploader
import browser_runtime
import captcha_ocr
from report_worker.executor import (
    RunExecutor,
    aggregate_project_results,
    limit_converted_file_rows,
    max_execution_retries,
    parse_converted_items,
    redact,
)
from report_worker.repository import Repository
from report_worker.storage import ArtifactStorage


def test_browser_profile_is_stable_per_site_and_account(monkeypatch, tmp_path):
    monkeypatch.setenv("REPORT_FORWARD_BROWSER_PROFILES", str(tmp_path))
    first = browser_runtime.browser_profile_dir("source", "same-account")
    second = browser_runtime.browser_profile_dir("source", "same-account")
    target = browser_runtime.browser_profile_dir("target", "same-account")

    assert first == second
    assert first != target
    assert "same-account" not in first


def test_close_driver_exports_session_cookies(monkeypatch, tmp_path):
    profile = tmp_path / "profile"
    profile.mkdir()

    class FakeDriver:
        _shanhuai_profile_dir = str(profile)
        quit_called = False

        def execute_cdp_cmd(self, command, _params):
            assert command == "Network.getAllCookies"
            return {
                "cookies": [{
                    "name": "auth", "value": "session", "domain": "example.test",
                    "path": "/", "session": True, "size": 11,
                }]
            }

        def quit(self):
            self.quit_called = True

    driver = FakeDriver()
    browser_runtime.close_driver(driver)

    stored = json.loads((profile / "session-cookies.json").read_text())
    assert stored == [{
        "name": "auth", "value": "session", "domain": "example.test", "path": "/",
    }]
    assert driver.quit_called is True


def test_ocr_consensus_accepts_two_matching_models(monkeypatch):
    class FakeModel:
        def __init__(self, text, confidence):
            self.text = text
            self.confidence = confidence

        def classification(self, _image, probability=False):
            return {"text": self.text, "confidence": self.confidence}

    models = iter([FakeModel("a1b2", 0.7), FakeModel("a1b2", 0.72)])
    monkeypatch.setattr(captcha_ocr, "_ocr_model", lambda *_args: next(models))
    monkeypatch.setattr(captcha_ocr, "_image_variants", lambda image: [("original", image)])

    text, details = captcha_ocr.recognize_target_code(b"image")

    assert text == "a1b2"
    assert details["votes"] == 2


def test_ocr_classify_supports_probability_arrays():
    class FakeModel:
        def classification(self, _image, probability=False):
            assert probability is True
            return {"text": "a1b2", "probability": [[0.1, 0.8], [0.2, 0.6]]}

    assert captcha_ocr._classify(FakeModel(), b"image") == ("a1b2", 0.7)


def test_ocr_consensus_rejects_weak_disagreement(monkeypatch):
    class FakeModel:
        def __init__(self, text):
            self.text = text

        def classification(self, _image, probability=False):
            return {"text": self.text, "confidence": 0.7}

    models = iter([FakeModel("a1b2"), FakeModel("c3d4")])
    monkeypatch.setattr(captcha_ocr, "_ocr_model", lambda *_args: next(models))
    monkeypatch.setattr(captcha_ocr, "_image_variants", lambda image: [("original", image)])

    text, details = captcha_ocr.recognize_target_code(b"image")

    assert text is None
    assert details["reason"] == "low_consensus"


def test_math_ocr_limits_charset_and_normalizes_operator(monkeypatch):
    class FakeModel:
        def classification(self, _image, probability=False):
            return {"text": "7T8=?", "confidence": 0.9}

    monkeypatch.setattr(captcha_ocr, "_ocr_model", lambda *_args: FakeModel())
    monkeypatch.setattr(captcha_ocr, "_image_variants", lambda image: [("original", image)])

    text, details = captcha_ocr.recognize_math_expression(b"image")

    assert text == "7x8=?"
    assert details["votes"] == 2


def test_parse_converted_items(tmp_path):
    path = tmp_path / "converted.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "sheet1"
    sheet.append(["模板"])
    sheet.append(["姓名", "性别", "户籍", "证件", "号码", "手机号", "地址"])
    sheet.append(["张三", "男", "省外农村户口", "居民身份证", "330203199001011234", "13800000000", "宁波"])
    workbook.save(path)

    items = parse_converted_items(path)

    assert len(items) == 1
    assert items[0]["name"] == "张三"
    assert items[0]["fingerprint"] == hashlib.sha256(b"330203199001011234").hexdigest()


def test_converter_filters_by_latest_entry_days(monkeypatch, tmp_path):
    source = tmp_path / "测试项目工人花名册.xlsx"
    template = tmp_path / "template.xlsx"
    output = tmp_path / "output"
    output.mkdir()
    today = date.today()

    source_book = openpyxl.Workbook()
    source_sheet = source_book.active
    source_sheet.title = "花名册"
    source_sheet.append(["标题"])
    source_sheet.append(["项目"])
    source_sheet.append(["序号", "姓名", "工种", "性别", "班组", "身份证号码", "地址", "银行", "账号", "开户地", "电话", "工资", "进场时间", "最新进场时间"])
    source_sheet.append([1, "张三", "工人", "男", "一组", "330203199001011234", "宁波", "", "", "", "13800000000", "", "", today.strftime("%Y-%m-%d 08:00")])
    source_sheet.append([2, "李四", "工人", "男", "一组", "330203199001011235", "宁波", "", "", "", "13800000001", "", "", (today - timedelta(days=31)).strftime("%Y-%m-%d 08:00")])
    source_sheet.append([3, "王五", "工人", "男", "一组", "330203199001011236", "宁波", "", "", "", "13800000002", "", "", ""])
    source_book.save(source)

    template_book = openpyxl.Workbook()
    template_book.active.title = "sheet1"
    template_book.save(template)

    converted = converter.convert_file(source, template, output, latest_entry_days=30)
    result_book = openpyxl.load_workbook(converted, data_only=True)
    try:
        result_sheet = result_book["sheet1"]
        assert result_sheet.cell(3, 1).value == "张三"
        assert result_sheet.cell(4, 1).value is None
    finally:
        result_book.close()


def test_downloaded_source_file_is_filtered_in_place(tmp_path):
    source = tmp_path / "测试项目工人花名册.xlsx"
    today = date.today()
    source_book = openpyxl.Workbook()
    source_sheet = source_book.active
    source_sheet.title = "花名册"
    source_sheet.append(["标题"])
    source_sheet.append(["项目"])
    source_sheet.append(["序号", "姓名", "工种", "性别", "班组", "身份证号码", "地址", "银行", "账号", "开户地", "电话", "工资", "进场时间", "最新进场时间"])
    source_sheet.append([1, "张三", "", "", "", "", "", "", "", "", "", "", "", today.strftime("%Y-%m-%d %H:%M")])
    source_sheet.append([2, "李四", "", "", "", "", "", "", "", "", "", "", "", (today - timedelta(days=31)).strftime("%Y-%m-%d %H:%M")])
    source_sheet.append([3, "王五", "", "", "", "", "", "", "", "", "", "", "", ""])
    source_book.save(source)

    stats = converter.filter_source_file(source, latest_entry_days=30)

    assert stats == {"retained_count": 1, "filtered_count": 2}
    filtered_book = openpyxl.load_workbook(source, data_only=True)
    try:
        filtered_sheet = filtered_book["花名册"]
        assert filtered_sheet.max_row == 4
        assert filtered_sheet.cell(4, 2).value == "张三"
    finally:
        filtered_book.close()


def test_test_upload_workbook_is_limited_to_one_person(tmp_path):
    path = tmp_path / "converted.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "sheet1"
    sheet.append(["模板"])
    sheet.append(["姓名", "性别", "户籍", "证件", "号码", "手机号", "地址"])
    sheet.append(["张三", "男", "", "", "330203199001011234", "", ""])
    sheet.append(["李四", "男", "", "", "330203199001011235", "", ""])
    workbook.save(path)

    retained = limit_converted_file_rows(path, 1)

    assert retained == 1
    assert [item["name"] for item in parse_converted_items(path)] == ["张三"]


def test_batch_result_does_not_collapse_when_one_error_says_existing():
    result = uploader.normalize_upload_result("177", "1", "部分人员已存在")

    assert result == {
        "total_rows": 177,
        "success_rows": 1,
        "failure_rows": 176,
        "already_exists": False,
        "person_details_available": False,
    }


def test_single_existing_person_is_counted_as_success():
    result = uploader.normalize_upload_result("1", "0", "该人员已存在")

    assert result["success_rows"] == 1
    assert result["failure_rows"] == 0
    assert result["already_exists"] is True


def test_error_workbook_extracts_hashed_person_results(tmp_path):
    path = tmp_path / "errors.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "sheet1"
    sheet.append(["姓名", "证件号码", "错误消息"])
    sheet.append(["ZPMC", "KJCGGBNR", "ErrorMsg"])
    sheet.append(["张三", "330203199001011234", "人员已存在"])
    sheet.append(["李四", "330203199001011235", "备案日期错误"])
    workbook.save(path)

    results = uploader.extract_error_person_results([path])

    assert results == [
        {
            "identity_fingerprint": hashlib.sha256(b"330203199001011234").hexdigest(),
            "error": "人员已存在",
        },
        {
            "identity_fingerprint": hashlib.sha256(b"330203199001011235").hexdigest(),
            "error": "备案日期错误",
        },
    ]


def test_error_workbook_falls_back_to_person_names(tmp_path):
    path = tmp_path / "name-errors.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(["姓名"])
    sheet.append(["ZPMC"])
    sheet.append(["张三"])
    sheet.append(["李四"])
    workbook.save(path)

    with ZipFile(path) as archive:
        entries = {name: archive.read(name) for name in archive.namelist()}
    sheet_xml = entries["xl/worksheets/sheet1.xml"]
    entries["xl/worksheets/sheet1.xml"] = sheet_xml.replace(
        b'<dimension ref="A1:A4" />', b'<dimension ref="A1:A1" />'
    )
    assert entries["xl/worksheets/sheet1.xml"] != sheet_xml
    with ZipFile(path, "w", ZIP_DEFLATED) as archive:
        for name, content in entries.items():
            archive.writestr(name, content)

    assert uploader.extract_error_person_results([path]) == [
        {"person_name": "张三", "error": "政府错误明细判定该人员失败"},
        {"person_name": "李四", "error": "政府错误明细判定该人员失败"},
    ]


def test_redact_sensitive_values():
    value = redact("手机号 13800000000 身份证 330203199001011234 验证码: 123456 code=654321 chat_id: oc_private")
    assert "13800000000" not in value
    assert "330203199001011234" not in value
    assert "123456" not in value
    assert "654321" not in value
    assert "oc_private" not in value
    assert "2026-07-23" in redact("验证码接收时间: 2026-07-23")


def test_target_login_propagates_failed_post_login_check(monkeypatch):
    login = target_login.TargetLogin({
        "credentials": {"target_site": {"username": "account", "password": "secret"}},
        "browser": {"headless": True},
    })
    for method in (
        "_init_driver", "_click_login_entry", "_switch_to_login_popup",
        "_click_legal_person_login", "_click_account_login", "_fill_credentials",
        "_fill_sms_code", "_confirm_login",
    ):
        monkeypatch.setattr(login, method, lambda *args, **kwargs: None)
    monkeypatch.setattr(login, "_solve_image_captcha", lambda: "abcd")
    monkeypatch.setattr(login, "_click_login_button", lambda: True)
    monkeypatch.setattr(login, "_check_captcha_error", lambda: False)
    monkeypatch.setattr(login, "_check_sms_popup_appeared", lambda: True)
    monkeypatch.setattr(login, "_wait_for_sms_code", lambda **kwargs: "123456")
    monkeypatch.setattr(login, "_check_sms_code_error", lambda: False)
    monkeypatch.setattr(login, "_check_login_success", lambda: False)
    monkeypatch.setattr(target_login.time, "sleep", lambda *_args: None)

    assert login.login() is False


def test_oss_storage_refuses_incomplete_credentials(monkeypatch):
    monkeypatch.setenv("STORAGE_DRIVER", "jdcloud_oss")
    for name in (
        "JD_OSS_BUCKET", "JD_OSS_ENDPOINT", "JD_OSS_ACCESS_KEY_ID",
        "JD_OSS_ACCESS_KEY_SECRET",
    ):
        monkeypatch.delenv(name, raising=False)

    try:
        ArtifactStorage()
    except RuntimeError as error:
        assert "JD_OSS_BUCKET" in str(error)
        assert "JD_OSS_ACCESS_KEY_SECRET" in str(error)
    else:
        raise AssertionError("incomplete OSS configuration must fail closed")


def test_oss_storage_uses_bucket_subdomain(monkeypatch):
    monkeypatch.setenv("STORAGE_DRIVER", "jdcloud_oss")
    monkeypatch.setenv("JD_OSS_BUCKET", "shanhuai-gc")
    monkeypatch.setenv("JD_OSS_ENDPOINT", "s3.cn-east-2.jdcloud-oss.com")
    monkeypatch.setenv("JD_OSS_ACCESS_KEY_ID", "test-key")
    monkeypatch.setenv("JD_OSS_ACCESS_KEY_SECRET", "test-secret")

    storage = ArtifactStorage()
    url = storage.client.generate_presigned_url(
        "put_object",
        Params={"Bucket": storage.bucket, "Key": "report-forward/test.xlsx"},
    )

    assert url.startswith("https://shanhuai-gc.s3.cn-east-2.jdcloud-oss.com/")
    assert storage.client.meta.config.request_checksum_calculation == "when_required"
    assert storage.client.meta.config.response_checksum_validation == "when_required"


def test_project_filter_only_keeps_exact_configured_names():
    downloader = executor_module.Downloader.__new__(executor_module.Downloader)
    downloader.download_settings = {
        "include_projects": ["测试项目"],
        "exclude_projects": [],
    }

    assert downloader._filter_projects(["测试项目", "测试项目二期", "其他项目"]) == ["测试项目"]


def test_project_filter_normalizes_full_width_punctuation_without_fuzzy_matching():
    downloader = executor_module.Downloader.__new__(executor_module.Downloader)
    downloader.download_settings = {
        "include_projects": ["东部新城工程(二期)"],
        "exclude_projects": [],
    }

    assert downloader._filter_projects([
        "东部新城工程（二期）",
        "东部新城工程（二期）景观段",
    ]) == ["东部新城工程（二期）"]


def test_selected_projects_stop_pagination_after_all_are_processed():
    downloader = executor_module.Downloader.__new__(executor_module.Downloader)
    downloader.download_settings = {
        "include_projects": ["项目甲", "项目乙", "已排除项目"],
        "exclude_projects": ["已排除项目"],
    }

    assert downloader._selected_projects_complete({"项目甲"}) is False
    assert downloader._selected_projects_complete({"项目甲", "项目乙"}) is True


def test_selected_projects_completion_normalizes_full_width_punctuation():
    downloader = executor_module.Downloader.__new__(executor_module.Downloader)
    downloader.download_settings = {
        "include_projects": ["东部新城工程(二期)"],
        "exclude_projects": [],
    }

    assert downloader._selected_projects_complete({"东部新城工程（二期）"}) is True


def test_login_feedback_does_not_wait_for_absent_optional_messages():
    class FakeDriver:
        def __init__(self):
            self.waits = []

        def implicitly_wait(self, seconds):
            self.waits.append(seconds)

        def find_elements(self, *_args):
            return []

    downloader = executor_module.Downloader.__new__(executor_module.Downloader)
    downloader.driver = FakeDriver()

    assert downloader._login_feedback() == ""
    assert downloader.driver.waits == [0, 5]


def test_source_entry_navigates_directly_to_project_list(monkeypatch):
    instance = executor_module.Downloader.__new__(executor_module.Downloader)
    class FakeDriver:
        current_url = "http://tg.91jtg.com/#/home"
        navigated = None
        def get(self, url):
            self.navigated = url
            self.current_url = url
    instance.driver = FakeDriver()
    monkeypatch.setattr(downloader.time, "sleep", lambda _seconds: None)

    assert instance._enter_project_list() is True
    assert instance.driver.navigated == "http://tg.91jtg.com/#/project/index"


def test_upload_with_row_failures_is_partial_success(monkeypatch, tmp_path):
    class FakeUploader:
        def __init__(self, _config):
            pass

        def run(self):
            return [{
                "project_name": "测试项目", "status": "validated",
                "total_rows": 10, "success_rows": 8, "failure_rows": 2,
            }]

    class FakeRepository:
        def __init__(self):
            self.project_update = None
            self.item_status = None

        def update_project(self, _project_id, **fields):
            self.project_update = fields

        def mark_project_items(self, _project_id, status, *_args):
            self.item_status = status

        def upsert_project(self, *_args):
            return "project-id"

    monkeypatch.setattr(executor_module, "Uploader", FakeUploader)
    executor = RunExecutor.__new__(RunExecutor)
    executor.config = {"browser": {"error_dir": str(tmp_path / "errors")}}
    executor.config_id = "config-id"
    executor.run_id = "run-id"
    executor.project_ids = {"测试项目": "project-id"}
    executor.successful_projects = 0
    executor.failed_projects = 0
    executor.repo = FakeRepository()
    executor.stage = lambda *_args, **_kwargs: None

    executor._upload()

    assert executor.successful_projects == 1
    assert executor.failed_projects == 1
    assert executor.repo.project_update["status"] == "partial_success"
    assert executor.repo.item_status == "result_unknown"


def test_upload_maps_error_detail_people_and_infers_success(monkeypatch, tmp_path):
    failed_fingerprint = hashlib.sha256(b"330203199001011234").hexdigest()

    class FakeUploader:
        def __init__(self, _config):
            pass

        def run(self):
            return [{
                "project_name": "测试项目", "status": "success",
                "total_rows": 2, "success_rows": 1, "failure_rows": 1,
                "person_details_available": True,
                "person_results": [{
                    "identity_fingerprint": failed_fingerprint,
                    "error": "身份证 330203199001011234 校验失败",
                }],
            }]

    class FakeRepository:
        def __init__(self):
            self.project_update = None
            self.item_result_args = None

        def update_project(self, _project_id, **fields):
            self.project_update = fields

        def mark_project_item_results(self, project_id, default_status, receipt, results):
            self.item_result_args = (project_id, default_status, receipt, results)

        def mark_project_items(self, *_args):
            raise AssertionError("complete error details must use per-person mapping")

        def upsert_project(self, *_args):
            return "project-id"

    monkeypatch.setattr(executor_module, "Uploader", FakeUploader)
    executor = RunExecutor.__new__(RunExecutor)
    executor.config = {"browser": {"error_dir": str(tmp_path / "errors")}}
    executor.config_id = "config-id"
    executor.run_id = "run-id"
    executor.project_ids = {"测试项目": "project-id"}
    executor.successful_projects = 0
    executor.failed_projects = 0
    executor.repo = FakeRepository()
    executor.stage = lambda *_args, **_kwargs: None

    executor._upload()

    assert executor.repo.project_update["status"] == "partial_success"
    project_id, default_status, receipt, results = executor.repo.item_result_args
    assert project_id == "project-id"
    assert default_status == "submitted"
    assert "person_results" not in receipt
    assert results == [{
        "identity_fingerprint": failed_fingerprint,
        "person_name": None,
        "error": "身份证 330203********1234 校验失败",
    }]


def test_max_execution_retries_defaults_to_three_and_is_capped(monkeypatch):
    monkeypatch.delenv("REPORT_FORWARD_MAX_RETRIES", raising=False)
    assert max_execution_retries() == 3
    monkeypatch.setenv("REPORT_FORWARD_MAX_RETRIES", "99")
    assert max_execution_retries() == 3
    monkeypatch.setenv("REPORT_FORWARD_MAX_RETRIES", "1")
    assert max_execution_retries() == 1
    monkeypatch.setenv("REPORT_FORWARD_MAX_RETRIES", "invalid")
    assert max_execution_retries() == 3


def test_executor_requeues_mid_run_exception(monkeypatch):
    class FakeRepository:
        def __init__(self):
            self.retry = None

        def event(self, *_args, **_kwargs):
            pass

        def schedule_retry(self, run_id, error, max_retries):
            self.retry = (run_id, error, max_retries)
            return True

        def complete(self, *_args):
            raise AssertionError("a retryable run must not be completed as failed")

    class FakeTemp:
        def cleanup(self):
            pass

    monkeypatch.setenv("REPORT_FORWARD_MAX_RETRIES", "3")
    executor = RunExecutor.__new__(RunExecutor)
    executor.repo = FakeRepository()
    executor.run_id = "run-id"
    executor.mode = "production"
    executor.context = {"stage": "download", "project_id": None}
    executor.temp = FakeTemp()
    executor._download = lambda: (_ for _ in ()).throw(RuntimeError("network error"))

    assert executor.execute() == "retrying"
    assert executor.repo.retry == ("run-id", "network error", 3)


def test_executor_finishes_failed_after_retry_limit(monkeypatch):
    class FakeRepository:
        def __init__(self):
            self.completed = None

        def event(self, *_args, **_kwargs):
            pass

        def schedule_retry(self, *_args):
            return False

        def complete(self, run_id, status, error):
            self.completed = (run_id, status, error)

    class FakeTemp:
        def cleanup(self):
            pass

    executor = RunExecutor.__new__(RunExecutor)
    executor.repo = FakeRepository()
    executor.run_id = "run-id"
    executor.mode = "production"
    executor.context = {"stage": "download", "project_id": None}
    executor.temp = FakeTemp()
    executor._download = lambda: (_ for _ in ()).throw(RuntimeError("still failing"))

    assert executor.execute() == "failed"
    assert executor.repo.completed == ("run-id", "failed", "still failing")


def test_repository_only_schedules_three_retries():
    class FakeResult:
        def __init__(self, row=None):
            self.row = row

        def fetchone(self):
            return self.row

    class FakeConnection:
        def __init__(self, attempt_count):
            self.attempt_count = attempt_count
            self.calls = []

        def transaction(self):
            return nullcontext()

        def execute(self, sql, params):
            self.calls.append((sql, params))
            if "SELECT attempt_count" in sql:
                return FakeResult({
                    "attempt_count": self.attempt_count,
                    "cancel_requested": False,
                })
            return FakeResult()

    def repository_for(attempt_count):
        repo = Repository.__new__(Repository)
        connection = FakeConnection(attempt_count)

        @contextmanager
        def open_connection():
            yield connection

        repo.connection = open_connection
        return repo, connection

    repo, connection = repository_for(3)
    assert repo.schedule_retry("run-id", "temporary error", 3) is True
    assert any("status='pending'" in sql for sql, _params in connection.calls)

    repo, connection = repository_for(4)
    assert repo.schedule_retry("run-id", "last error", 3) is False
    assert not any("status='pending'" in sql for sql, _params in connection.calls)


def test_uploader_retries_only_the_failed_project(monkeypatch, tmp_path):
    output_root = tmp_path / "output"
    dated_output = output_root / executor_module.datetime.now().strftime("%Y%m%d")
    dated_output.mkdir(parents=True)
    (dated_output / "测试项目.xlsx").touch()
    config = {
        "browser": {
            "output_dir": str(output_root),
            "error_dir": str(tmp_path / "errors"),
            "upload_timeout": 1,
        },
        "runtime": {"max_execution_retries": 3},
    }
    instance = uploader.Uploader(config, driver=object())
    attempts = []
    restarts = []

    def upload_file(_path, project_name, is_first=False):
        attempts.append((project_name, is_first))
        if len(attempts) == 1:
            raise RuntimeError("browser disconnected")
        return {"project_name": project_name, "status": "success"}

    monkeypatch.setattr(instance, "_upload_single_file", upload_file)
    monkeypatch.setattr(uploader, "split_upload_workbook", lambda path: [path])
    monkeypatch.setattr(instance, "_log_page_state", lambda *_args: None)
    monkeypatch.setattr(instance, "_restart_session", lambda: restarts.append(True))

    results = instance._do_uploads()

    assert len(attempts) == 2
    assert attempts[1][1] is True
    assert restarts == [True]


def test_split_upload_workbook_uses_200_row_batches(tmp_path):
    path = tmp_path / "20260803_测试项目_姜太公导出.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "sheet1"
    for row_number in range(3, 656):
        sheet.cell(row=row_number, column=1, value=f"人员{row_number}")
        sheet.cell(row=row_number, column=5, value=f"证件{row_number}")
    workbook.save(path)
    workbook.close()

    batches = uploader.split_upload_workbook(str(path))

    assert [uploader.count_upload_rows(batch) for batch in batches] == [200, 200, 200, 53]
    assert all(uploader.extract_project_name(batch) == "测试项目" for batch in batches)


def test_aggregate_project_results_sums_upload_batches():
    results = aggregate_project_results([
        {
            "project_name": "测试项目", "status": "success",
            "total_rows": 200, "success_rows": 190, "failure_rows": 10,
            "person_details_available": True, "person_results": [{"person_name": "甲"}],
        },
        {
            "project_name": "测试项目", "status": "success",
            "total_rows": 53, "success_rows": 50, "failure_rows": 3,
            "person_details_available": True, "person_results": [{"person_name": "乙"}],
        },
    ])

    assert len(results) == 1
    assert results[0]["total_rows"] == 253
    assert results[0]["success_rows"] == 240
    assert results[0]["failure_rows"] == 13
    assert len(results[0]["person_results"]) == 2
    assert len(results[0]["batches"]) == 2
    assert results[0]["status"] == "success"


def test_aggregate_project_results_does_not_spread_single_existing_person():
    results = aggregate_project_results([
        {
            "project_name": "测试项目", "status": "success",
            "total_rows": 200, "success_rows": 200, "failure_rows": 0,
        },
        {
            "project_name": "测试项目", "status": "success",
            "total_rows": 1, "success_rows": 1, "failure_rows": 0,
            "already_exists": True,
        },
    ])

    assert results[0]["total_rows"] == 201
    assert results[0]["already_exists"] is False

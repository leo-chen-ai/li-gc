import hashlib
import json
import logging
import os
import re
import shutil
import sys
import tempfile
from datetime import datetime
from pathlib import Path

import openpyxl
import yaml

from .storage import ArtifactStorage


ROOT = Path(__file__).resolve().parents[1]
LEGACY = ROOT / "legacy"
ASSETS = ROOT / "assets"
if str(LEGACY) not in sys.path:
    sys.path.insert(0, str(LEGACY))

from converter import _extract_project_name, batch_convert, filter_source_file  # noqa: E402
from downloader import Downloader  # noqa: E402
from target_login import TargetLogin  # noqa: E402
from uploader import Uploader, extract_project_name  # noqa: E402


class CancelledError(RuntimeError):
    pass


def max_execution_retries():
    try:
        configured = int(os.environ.get("REPORT_FORWARD_MAX_RETRIES", "3"))
    except ValueError:
        configured = 3
    return min(max(configured, 0), 3)


def redact(value):
    value = str(value)
    value = re.sub(r"(?<!\d)1\d{10}(?!\d)", lambda m: m.group(0)[:3] + "****" + m.group(0)[-4:], value)
    value = re.sub(r"(?<!\d)\d{17}[0-9Xx](?!\d)", lambda m: m.group(0)[:6] + "********" + m.group(0)[-4:], value)
    value = re.sub(r"(?i)(验证码\s*(?:为|[:：=])\s*)\d{4,6}\b", r"\1******", value)
    value = re.sub(r"(?i)(code\s*[=:]\s*)\d{4,6}\b", r"\1******", value)
    value = re.sub(r"(?i)(chat_id\s*[=:]\s*)[^,\s)]+", r"\1******", value)
    value = re.sub(r"\boc_[0-9A-Za-z]+\b", "oc_******", value)
    return value


class DatabaseLogHandler(logging.Handler):
    def __init__(self, repository, run_id, context):
        super().__init__(logging.INFO)
        self.repository = repository
        self.run_id = run_id
        self.context = context

    def emit(self, record):
        try:
            level = "warning" if record.levelno == logging.WARNING else "error" if record.levelno >= logging.ERROR else "info"
            self.repository.event(
                self.run_id, self.context["stage"], redact(record.getMessage()), level,
                self.context.get("project_id"),
            )
        except Exception:
            pass


class RedactingFilter(logging.Filter):
    def filter(self, record):
        record.msg = redact(record.getMessage())
        record.args = ()
        return True


class RunExecutor:
    def __init__(self, repository, run):
        self.repo = repository
        self.run = run
        self.run_id = run["id"]
        self.config_id = run["config_id"]
        self.mode = run["run_mode"]
        self.options = run.get("options") or {}
        self.config_row = repository.runtime_config(self.config_id)
        if not self.config_row:
            raise RuntimeError("报送配置不存在或已删除")
        self.storage = ArtifactStorage()
        self.context = {"stage": "starting", "project_id": None}
        self.temp = tempfile.TemporaryDirectory(prefix=f"report-{self.run_id}-")
        self.work_dir = Path(self.temp.name)
        self.config = self._legacy_config()
        self.project_ids = {}
        self.successful_projects = 0
        self.failed_projects = 0
        self._configure_logging()

    def _configure_logging(self):
        handler = DatabaseLogHandler(self.repo, self.run_id, self.context)
        stdout = logging.StreamHandler(sys.stdout)
        stdout.addFilter(RedactingFilter())
        logging.basicConfig(
            level=logging.INFO,
            format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
            handlers=[stdout, handler],
            force=True,
        )

    def _legacy_config(self):
        download_root = self.work_dir / "download"
        output_root = self.work_dir / "output"
        error_root = self.work_dir / "error"
        verification = self.config_row.get("verification_config") or {}
        include = "all" if self.config_row["project_mode"] == "all" else list(self.config_row["include_projects"])
        config = {
            "credentials": {
                "source_site": {"username": self.config_row["source_username"], "password": self.config_row["source_password"]},
                "target_site": {"username": self.config_row["target_username"], "password": self.config_row["target_password"]},
            },
            "download": {
                "include_projects": include,
                "exclude_projects": list(self.config_row["exclude_projects"]),
                "latest_entry_days": int((self.config_row.get("settings") or {}).get("latest_entry_days", 1)),
            },
            "browser": {
                "download_dir": str(download_root), "output_dir": str(output_root), "error_dir": str(error_root),
                "diagnostics_dir": str(self.work_dir / "diagnostics"),
                "template_file": str(ASSETS / "建筑项目人员备案信息模板.xlsx"),
                "headless": bool((self.config_row.get("settings") or {}).get("headless", True)),
                "upload_timeout": int((self.config_row.get("settings") or {}).get("upload_timeout_minutes", 10)),
            },
            "feishu": verification,
            "email": {"enabled": False},
            "runtime": {
                "validate_only": self.mode == "test_upload_validate",
                "max_execution_retries": max_execution_retries(),
            },
        }
        config_path = self.work_dir / "runtime-config.yaml"
        config_path.write_text(yaml.safe_dump(config, allow_unicode=True), encoding="utf-8")
        os.environ["REPORT_FORWARD_CONFIG_PATH"] = str(config_path)
        os.environ["REPORT_FORWARD_CODES_CSV"] = str(self.work_dir / "verification_codes.csv")
        os.environ["REPORT_FORWARD_PROCESSED_IDS"] = str(self.work_dir / "processed_ids.txt")
        return config

    def stage(self, value, message=None):
        self.context["stage"] = value
        self.context["project_id"] = None
        self.repo.set_stage(self.run_id, value)
        if message:
            self.repo.event(self.run_id, value, message)
        self.check_cancelled()

    def check_cancelled(self):
        if self.repo.cancelled(self.run_id):
            raise CancelledError("任务已由管理员取消")

    def execute(self):
        try:
            self.repo.event(self.run_id, "starting", f"开始执行 {self.mode}")
            if self.mode == "test_source_login":
                self._test_source_login()
            elif self.mode == "test_project_list":
                self._test_project_list()
            elif self.mode == "test_target_login":
                self._test_target_login()
            elif self.mode in {"test_upload_validate", "test_submit"}:
                self._prepare_converted_from_run()
                self._upload()
            elif self.mode == "test_transform":
                if self.options.get("source_run_id"):
                    self._prepare_source_from_run()
                else:
                    self._download()
                self._convert()
            else:
                self._download()
                if self.mode != "test_download":
                    self._convert()
                    self._upload()
            self.stage("finalizing", "正在汇总运行结果")
            # 对方平台拒绝、数据已存在或业务规则跳过，都表示自动化流程已经
            # 正常执行完毕。只有抛出系统执行异常时，任务才记为 failed。
            status = "success"
            self.repo.complete(self.run_id, status)
            self.repo.event(
                self.run_id,
                "completed",
                f"任务执行完成：成功项目 {self.successful_projects} 个，含跳过项目 {self.failed_projects} 个",
            )
            return status
        except CancelledError as error:
            self.repo.event(self.run_id, self.context["stage"], str(error), "warning")
            self.repo.complete(self.run_id, "cancelled", str(error))
            return "cancelled"
        except Exception as error:
            logging.exception("任务执行失败")
            safe_error = redact(error)
            if self.repo.schedule_retry(
                self.run_id, safe_error, max_execution_retries()
            ):
                return "retrying"
            self.repo.complete(self.run_id, "failed", safe_error)
            return "failed"
        finally:
            self.temp.cleanup()

    def _test_source_login(self):
        self.stage("source_login", "测试源网站登录")
        downloader = Downloader(self.config)
        try:
            if not downloader.login():
                raise RuntimeError("源网站登录测试失败")
        finally:
            downloader.close()
        self.repo.event(self.run_id, "source_login", "源网站登录测试通过")

    def _test_project_list(self):
        self.stage("project_list", "测试读取源网站项目列表")
        names = Downloader(self.config).discover_projects()
        if not names:
            raise RuntimeError("未读取到任何项目")
        for name in names:
            self.project_ids[name] = self.repo.upsert_project(self.run_id, name)
        self.repo.event(self.run_id, "project_list", f"读取到 {len(names)} 个项目")

    def _test_target_login(self):
        self.stage("target_login", "测试目标网站登录和二次验证码")
        if not TargetLogin(self.config).run():
            raise RuntimeError("目标网站登录测试失败")
        self.repo.event(self.run_id, "target_login", "目标网站登录测试通过")

    def _download(self):
        self.stage("download", "开始从源网站下载项目花名册")
        files = Downloader(self.config).run()
        if not files:
            raise RuntimeError("源网站没有下载到文件")
        for file_path in files:
            self.check_cancelled()
            project_name = _extract_project_name(Path(file_path).name)
            project_id = self.repo.upsert_project(self.run_id, project_name, "downloaded", "download")
            self.project_ids[project_name] = project_id
            latest_entry_days = self.config["download"]["latest_entry_days"]
            filter_stats = filter_source_file(file_path, latest_entry_days)
            stored = self.storage.put(self.config_id, self.run_id, "source", file_path)
            self.repo.add_artifact(self.run_id, project_id, "source", stored)
            self.repo.update_project(
                project_id,
                status="downloaded",
                current_stage="download",
                source_row_count=filter_stats["retained_count"],
            )
        self.repo.event(self.run_id, "download", f"下载完成，共 {len(files)} 个文件")

    def _prepare_source_from_run(self):
        self.stage("prepare_source", "读取已留存的源文件")
        artifacts = self.repo.artifacts(self.options["source_run_id"], self.config_id, "source")
        if not artifacts:
            raise RuntimeError("来源任务没有原始文件")
        date_dir = Path(self.config["browser"]["download_dir"]) / datetime.now().strftime("%Y%m%d")
        date_dir.mkdir(parents=True, exist_ok=True)
        for artifact in artifacts:
            destination = date_dir / artifact["original_filename"]
            self.storage.get(artifact["object_key"], destination)
            name = artifact.get("external_project_name") or _extract_project_name(destination.name)
            self.project_ids[name] = self.repo.upsert_project(self.run_id, name, "downloaded", "prepare_source")

    def _convert(self):
        self.stage("transform", "开始转换花名册")
        files = batch_convert(self.config)
        if not files:
            raise RuntimeError("没有生成转换文件")
        for file_path in files:
            self.check_cancelled()
            project_name = extract_project_name(file_path)
            project_id = self.project_ids.get(project_name) or self.repo.upsert_project(self.run_id, project_name)
            self.project_ids[project_name] = project_id
            if self.mode == "test_full":
                retained = limit_converted_file_rows(file_path, 1)
                self.repo.event(
                    self.run_id,
                    "transform",
                    f"全流程测试仅保留 {retained} 条人员数据，避免批量发送",
                )
            items = parse_converted_items(file_path)
            self.repo.add_items(self.run_id, project_id, items)
            stored = self.storage.put(self.config_id, self.run_id, "converted", file_path)
            self.repo.add_artifact(self.run_id, project_id, "converted", stored)
            self.repo.update_project(
                project_id, status="converted", current_stage="transform",
                converted_row_count=len(items), source_row_count=len(items),
            )
        self.repo.event(self.run_id, "transform", f"转换完成，共 {len(files)} 个项目")

    def _prepare_converted_from_run(self):
        self.stage("prepare_upload", "读取已留存的转换文件")
        artifacts = self.repo.converted_artifacts(self.options["source_run_id"], self.config_id)
        if not artifacts:
            raise RuntimeError("来源任务没有转换文件")
        output_dir = Path(self.config["browser"]["output_dir"]) / datetime.now().strftime("%Y%m%d")
        output_dir.mkdir(parents=True, exist_ok=True)
        for artifact in artifacts:
            destination = output_dir / artifact["original_filename"]
            self.storage.get(artifact["object_key"], destination)
            if self.mode in {"test_upload_validate", "test_submit"}:
                retained = limit_converted_file_rows(destination, 1)
                self.repo.event(
                    self.run_id,
                    "prepare_upload",
                    f"测试报送仅保留 {retained} 条人员数据，避免批量发送",
                )
            name = artifact.get("external_project_name") or extract_project_name(destination)
            project_id = self.repo.upsert_project(self.run_id, name, "converted", "prepare_upload")
            self.project_ids[name] = project_id
            items = parse_converted_items(destination)
            self.repo.add_items(self.run_id, project_id, items)
            stored = self.storage.put(self.config_id, self.run_id, "converted", destination)
            self.repo.add_artifact(self.run_id, project_id, "converted", stored)
            self.repo.update_project(project_id, converted_row_count=len(items), source_row_count=len(items))

    def _upload(self):
        self.stage("target_upload", "开始登录目标网站并上报")
        try:
            results = aggregate_project_results(Uploader(self.config).run())
        finally:
            self._store_diagnostics()
        if not results:
            raise RuntimeError("目标网站未产生上传结果")
        execution_errors = []
        for result in results:
            name = result["project_name"]
            project_id = self.project_ids.get(name) or self.repo.upsert_project(self.run_id, name)
            person_results = [
                {
                    "identity_fingerprint": item.get("identity_fingerprint"),
                    "person_name": item.get("person_name"),
                    "error": redact(item.get("error", "政府错误明细判定该人员失败")),
                }
                for item in result.get("person_results", [])
                if item.get("identity_fingerprint") or item.get("person_name")
            ]
            receipt = {key: value for key, value in result.items() if key != "person_results"}
            total = result.get("total_rows") or 0
            success = result.get("success_rows") or 0
            failure = result.get("failure_rows") if result.get("failure_rows") is not None else (1 if result["status"] == "failed" else 0)
            if result["status"] == "failed":
                self.failed_projects += 1
                self.repo.update_project(project_id, status="failed", current_stage="target_upload", last_error=redact(result.get("error", "上传失败")), upload_failure_count=failure)
                self.repo.mark_project_items(project_id, "failed", receipt, redact(result.get("error", "上传失败")))
                if result.get("execution_error"):
                    execution_errors.append(
                        f"{name}: {redact(result.get('error', '上传执行失败'))}"
                    )
            else:
                if success > 0 or failure == 0:
                    self.successful_projects += 1
                if failure:
                    self.failed_projects += 1
                project_status = "partial_success" if failure else "validated" if result["status"] == "validated" else "success"
                self.repo.update_project(
                    project_id, status=project_status, current_stage="target_upload", target_receipt=receipt,
                    last_error=redact(result.get("error", "部分批次上传失败")) if failure else None,
                    upload_total_count=total, upload_success_count=success, upload_failure_count=failure,
                )
                if total > 1 and failure and person_results:
                    default_status = (
                        "validated" if result["status"] == "validated" else "submitted"
                    ) if result.get("person_details_available") else "result_unknown"
                    self.repo.mark_project_item_results(
                        project_id, default_status, receipt, person_results
                    )
                else:
                    if total > 1 and failure and not result.get("person_details_available"):
                        item_status = "result_unknown"
                    else:
                        item_status = "validated_with_errors" if result["status"] == "validated" and failure else "validated" if result["status"] == "validated" else "submitted_with_errors" if failure else "submitted"
                    self.repo.mark_project_items(project_id, item_status, receipt)
        if execution_errors:
            raise RuntimeError("目标网站自动化执行失败：" + "；".join(execution_errors))
        error_dir = Path(self.config["browser"]["error_dir"]) / datetime.now().strftime("%Y%m%d")
        if error_dir.exists():
            for path in error_dir.glob("*.xlsx"):
                project_id = next((pid for name, pid in self.project_ids.items() if name[:6] in path.name), None)
                stored = self.storage.put(self.config_id, self.run_id, "error_detail", path)
                self.repo.add_artifact(self.run_id, project_id, "error_detail", stored)

    def _store_diagnostics(self):
        configured_dir = self.config["browser"].get("diagnostics_dir")
        if not configured_dir:
            return
        diagnostics_dir = Path(configured_dir)
        if not diagnostics_dir.exists():
            return
        artifact_types = {".png": "diagnostic_screenshot", ".html": "diagnostic_html"}
        for path in sorted(diagnostics_dir.iterdir()):
            artifact_type = artifact_types.get(path.suffix.lower())
            if not artifact_type:
                continue
            try:
                stored = self.storage.put(self.config_id, self.run_id, artifact_type, path)
                self.repo.add_artifact(self.run_id, None, artifact_type, stored)
            except Exception as error:
                logging.warning("诊断证据留存失败 %s: %s", path.name, error)


def aggregate_project_results(results):
    grouped = {}
    for result in results:
        project_name = result["project_name"]
        aggregate = grouped.setdefault(project_name, {
            "project_name": project_name,
            "status": "failed",
            "total_rows": 0,
            "success_rows": 0,
            "failure_rows": 0,
            "already_exists": False,
            "execution_error": False,
            "person_details_available": True,
            "person_results": [],
            "errors": [],
            "batches": [],
        })
        total = result.get("total_rows") or 0
        success = result.get("success_rows") or 0
        failure = result.get("failure_rows")
        failure = failure if failure is not None else (total if result.get("status") == "failed" else 0)
        aggregate["total_rows"] += total
        aggregate["success_rows"] += success
        aggregate["failure_rows"] += failure
        # “已存在”只可靠描述单行批次，不能扩散成整个项目的人员结果。
        if total == 1 and result.get("already_exists"):
            aggregate["already_exists"] = True
        if failure:
            aggregate["person_details_available"] = (
                aggregate["person_details_available"]
                and bool(result.get("person_details_available"))
            )
        aggregate["person_results"].extend(result.get("person_results") or [])
        if result.get("error") and result["error"] not in aggregate["errors"]:
            aggregate["errors"].append(result["error"])
        aggregate["execution_error"] = (
            aggregate["execution_error"] or bool(result.get("execution_error"))
        )
        aggregate["batches"].append({
            key: value for key, value in result.items()
            if key not in {"person_results", "project_name"}
        })
        if result.get("status") == "validated" and aggregate["status"] == "failed":
            aggregate["status"] = "validated"
        elif result.get("status") == "success":
            aggregate["status"] = "success"

    for aggregate in grouped.values():
        if aggregate["total_rows"] > 1:
            aggregate["already_exists"] = False
        if aggregate["errors"]:
            aggregate["error"] = "；".join(aggregate["errors"])
        aggregate.pop("errors", None)
    return list(grouped.values())


def parse_converted_items(file_path):
    workbook = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
    try:
        sheet = workbook["sheet1"]
        items = []
        for row_no, row in enumerate(sheet.iter_rows(min_row=3, values_only=True), start=3):
            name = str(row[0] or "").strip()
            identity = str(row[4] or "").strip().upper()
            if not name or not identity:
                continue
            items.append({
                "row_no": row_no, "name": name, "gender": str(row[1] or "").strip(),
                "household_type": str(row[2] or "").strip(), "identity_type": str(row[3] or "").strip(),
                "identity": identity, "phone": str(row[5] or "").strip(), "address": str(row[6] or "").strip(),
                "fingerprint": hashlib.sha256(identity.encode("utf-8")).hexdigest(),
            })
        return items
    finally:
        workbook.close()


def limit_converted_file_rows(file_path, max_rows=1):
    """Keep only the first N valid rows in the temporary workbook used for testing."""
    workbook = openpyxl.load_workbook(file_path)
    try:
        if "sheet1" not in workbook.sheetnames:
            raise RuntimeError("转换文件中缺少 sheet1 工作表")
        sheet = workbook["sheet1"]
        valid_rows = []
        for row_number in range(3, sheet.max_row + 1):
            name = str(sheet.cell(row=row_number, column=1).value or "").strip()
            identity = str(sheet.cell(row=row_number, column=5).value or "").strip()
            if name and identity:
                valid_rows.append(row_number)
        keep = set(valid_rows[:max(0, int(max_rows))])
        for row_number in reversed(valid_rows):
            if row_number not in keep:
                sheet.delete_rows(row_number, 1)
        workbook.save(file_path)
        return len(keep)
    finally:
        workbook.close()

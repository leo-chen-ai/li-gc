import openpyxl
import shutil
import os
import re
import logging
from datetime import date, datetime, timedelta

logger = logging.getLogger(__name__)


def filter_source_file(src_path, latest_entry_days=30):
    """Filter a downloaded roster in place by its latest-entry date."""
    wb_src = openpyxl.load_workbook(src_path)
    try:
        if '花名册' not in wb_src.sheetnames:
            raise RuntimeError("下载文件中缺少‘花名册’工作表")

        ws_src = wb_src['花名册']
        header_row = None
        latest_entry_column = None
        for row in ws_src.iter_rows(min_row=1, max_row=min(ws_src.max_row, 20)):
            for cell in row:
                if str(cell.value or '').strip() == '最新进场时间':
                    header_row = cell.row
                    latest_entry_column = cell.column
                    break
            if header_row is not None:
                break

        if header_row is None or latest_entry_column is None:
            raise RuntimeError("下载文件中未找到‘最新进场时间’列，已停止处理以避免报送全部人员")

        days = max(1, int(latest_entry_days))
        cutoff_date = date.today() - timedelta(days=days)
        retained_count = 0
        rows_to_delete = []
        for row_number in range(header_row + 1, ws_src.max_row + 1):
            latest_entry = ws_src.cell(row=row_number, column=latest_entry_column).value
            latest_entry_date = _parse_date(latest_entry)
            if latest_entry_date is None or latest_entry_date < cutoff_date:
                rows_to_delete.append(row_number)
            else:
                retained_count += 1

        # Delete contiguous ranges from bottom to top so row numbers remain stable.
        delete_ranges = []
        for row_number in rows_to_delete:
            if delete_ranges and row_number == delete_ranges[-1][1] + 1:
                delete_ranges[-1] = (delete_ranges[-1][0], row_number)
            else:
                delete_ranges.append((row_number, row_number))
        for start, end in reversed(delete_ranges):
            ws_src.delete_rows(start, end - start + 1)

        wb_src.save(src_path)
        filtered_count = len(rows_to_delete)
        logger.info(
            "下载花名册已按最新进场时间过滤：最近 %s 天，保留 %s 条，过滤 %s 条",
            days, retained_count, filtered_count,
        )
        return {"retained_count": retained_count, "filtered_count": filtered_count}
    finally:
        wb_src.close()


def batch_convert(config):
    template_file = os.path.abspath(config['browser']['template_file'])
    download_dir = os.path.join(
        os.path.abspath(config['browser']['download_dir']),
        datetime.now().strftime('%Y%m%d')
    )
    output_dir = os.path.join(
        os.path.abspath(config['browser']['output_dir']),
        datetime.now().strftime('%Y%m%d')
    )

    os.makedirs(output_dir, exist_ok=True)

    if not os.path.exists(download_dir):
        logger.error(f"下载目录不存在: {download_dir}")
        return []

    xlsx_files = [
        f for f in os.listdir(download_dir)
        if f.endswith('.xlsx') and not f.startswith('~$')
    ]

    if not xlsx_files:
        logger.warning("下载目录中没有找到 .xlsx 文件")
        return []

    results = []
    for filename in xlsx_files:
        src_path = os.path.join(download_dir, filename)
        try:
            latest_entry_days = int(config.get('download', {}).get('latest_entry_days', 30))
            out_path = convert_file(src_path, template_file, output_dir, latest_entry_days)
            if out_path:
                results.append(out_path)
                logger.info(f"转换成功: {filename} -> {os.path.basename(out_path)}")
        except Exception as e:
            logger.error(f"转换失败 {filename}: {e}")

    logger.info(f"批量转换完成，共 {len(results)}/{len(xlsx_files)} 个文件成功")
    return results


def convert_file(src_path, template_path, output_dir, latest_entry_days=30):
    filename = os.path.basename(src_path)
    date_str = datetime.now().strftime('%Y%m%d')

    project_name = _extract_project_name(filename)
    if not project_name:
        project_name = os.path.splitext(filename)[0]

    out_filename = f"{date_str}_{project_name}_姜太公导出.xlsx"
    out_path = os.path.join(output_dir, out_filename)

    shutil.copy2(template_path, out_path)

    wb_src = openpyxl.load_workbook(src_path)
    ws_src = wb_src['花名册']

    wb_out = openpyxl.load_workbook(out_path)
    ws_out = wb_out['sheet1']

    row_out = 3
    count = 0
    filtered_count = 0
    cutoff_date = date.today() - timedelta(days=max(1, int(latest_entry_days)))

    for row in ws_src.iter_rows(min_row=4, max_row=ws_src.max_row, values_only=False):
        name = row[1].value if len(row) > 1 else None
        id_number = row[5].value if len(row) > 5 else None
        latest_entry = row[13].value if len(row) > 13 else None

        if not name or not re.match(r'^[\u4e00-\u9fff]+$', str(name)):
            continue
        if not id_number:
            continue
        latest_entry_date = _parse_date(latest_entry)
        if latest_entry_date is None or latest_entry_date < cutoff_date:
            filtered_count += 1
            continue

        gender = row[3].value if len(row) > 3 else ''
        address = row[6].value if len(row) > 6 else ''
        phone = row[10].value if len(row) > 10 else ''

        ws_out.cell(row=row_out, column=1, value=name)
        ws_out.cell(row=row_out, column=2, value=gender)
        ws_out.cell(row=row_out, column=3, value='省外农村户口')
        ws_out.cell(row=row_out, column=4, value='居民身份证')
        ws_out.cell(row=row_out, column=5, value=id_number)
        ws_out.cell(row=row_out, column=6, value=phone)
        ws_out.cell(row=row_out, column=7, value=address)
        ws_out.cell(row=row_out, column=9, value='是')

        row_out += 1
        count += 1

    wb_out.save(out_path)
    wb_src.close()
    wb_out.close()

    logger.info(
        "最新进场时间保留最近 %s 天：写入 %s 条，过滤 %s 条 -> %s",
        latest_entry_days, count, filtered_count, out_filename,
    )
    return out_path


def _parse_date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value or '').strip()
    if not text:
        return None
    for fmt in ('%Y-%m-%d %H:%M', '%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%Y/%m/%d %H:%M', '%Y/%m/%d'):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _extract_project_name(filename):
    name = filename
    for suffix in ['项目工人花名册', '工人花名册', '花名册']:
        if suffix in name:
            name = name.split(suffix)[0]
            break
    name = name.rstrip('的')
    name = re.sub(r'[\(\)（）]', '', name)
    name = name.strip('_').strip('-').strip()
    return name if name else os.path.splitext(filename)[0]

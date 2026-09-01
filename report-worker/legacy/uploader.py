import os
import re
import time
import shutil
import subprocess
import logging
import hashlib
from datetime import datetime

import openpyxl

from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import (
    TimeoutException,
    NoSuchElementException,
    ElementClickInterceptedException,
)
from browser_runtime import browser_profile_dir, close_driver, create_driver

logger = logging.getLogger(__name__)

HOME_URL = "https://www.zjzwfw.gov.cn/zjservice-fe/#/home"
DEFAULT_UPLOAD_TIMEOUT = 900
PAGE_LOAD_TIMEOUT = 30
EXISTING_MARKERS = ('已存在', '已经存在', '人员已备案', '重复参保', '重复申报')


class TargetSessionExpired(RuntimeError):
    """The public home page looks logged in, but the user center rejected it."""


def normalize_upload_result(total_rows, success_rows, result_text):
    total = int(total_rows) if str(total_rows).isdigit() else None
    success = int(success_rows) if str(success_rows).isdigit() else None
    # “已存在”只能明确对应到单行测试。批量文件里只要某一条出现
    # 该文案就把整批改成 1/0，会丢失政府返回的真实汇总。
    already_exists = (
        total == 1
        and success == 0
        and any(marker in result_text for marker in EXISTING_MARKERS)
    )
    if already_exists:
        success = 1
    failure = total - success if total is not None and success is not None else None
    return {
        'total_rows': total,
        'success_rows': success,
        'failure_rows': failure,
        'already_exists': already_exists,
        'person_details_available': total == 1,
    }


def extract_error_person_results(paths):
    """Read government error workbooks without retaining plaintext identities."""
    results = []
    seen_fingerprints = set()
    identity_headers = {'证件号码', '身份证号', '身份证号码', 'KJCGGBNR'}
    name_headers = {'姓名', '人员姓名', 'ZPMC'}
    error_headers = {'错误消息', '错误原因', '失败原因', 'ErrorMsg'}

    for path in paths:
        try:
            # Government workbooks can advertise a stale A1:A1 worksheet dimension.
            # Normal mode discovers the populated cells instead of trusting that metadata.
            workbook = openpyxl.load_workbook(path, data_only=True, read_only=False)
            try:
                for sheet in workbook.worksheets:
                    rows = list(sheet.iter_rows(values_only=True))
                    if not rows:
                        continue
                    header_rows = rows[:2]
                    identity_column = next(
                        (
                            index
                            for headers in header_rows
                            for index, value in enumerate(headers)
                            if str(value or '').strip() in identity_headers
                        ),
                        None,
                    )
                    error_column = next(
                        (
                            index
                            for headers in header_rows
                            for index, value in enumerate(headers)
                            if str(value or '').strip() in error_headers
                        ),
                        None,
                    )
                    name_column = next(
                        (
                            index
                            for headers in header_rows
                            for index, value in enumerate(headers)
                            if str(value or '').strip() in name_headers
                        ),
                        None,
                    )
                    if identity_column is None and name_column is None:
                        continue
                    for row in rows[2:]:
                        identity = (
                            str(row[identity_column] or '').strip().upper()
                            if identity_column is not None and identity_column < len(row)
                            else ''
                        )
                        person_name = (
                            str(row[name_column] or '').strip()
                            if name_column is not None and name_column < len(row)
                            else ''
                        )
                        if not identity and not person_name:
                            continue
                        fingerprint = (
                            hashlib.sha256(identity.encode('utf-8')).hexdigest()
                            if identity else None
                        )
                        if fingerprint and fingerprint in seen_fingerprints:
                            continue
                        error = (
                            str(row[error_column] or '').strip()
                            if error_column is not None and error_column < len(row)
                            else ''
                        )
                        result = {
                            'error': error or '政府错误明细判定该人员失败',
                        }
                        if fingerprint:
                            result['identity_fingerprint'] = fingerprint
                            seen_fingerprints.add(fingerprint)
                        else:
                            result['person_name'] = person_name
                        results.append(result)
            finally:
                workbook.close()
        except Exception as error:
            logger.warning(f"解析错误明细人员失败: {error}")

    return results


def _find_first(driver, selectors, check_displayed=True):
    for by, value in selectors:
        els = driver.find_elements(by, value)
        for el in els:
            try:
                if not check_displayed or el.is_displayed():
                    return el
            except Exception:
                continue
    return None


def extract_project_name(file_path):
    filename = os.path.basename(file_path)
    name = filename
    for suffix in ['_姜太公导出.xlsx', '_姜太公导出', '项目工人花名册', '工人花名册']:
        if suffix in name:
            name = name.split(suffix)[0]
            break
    name = name.split('_', 1)[-1] if '_' in name else name
    name = re.sub(r'_第\d+批$', '', name)
    name = re.sub(r'[\(\)（）]', '', name)
    name = name.strip('_').strip('-').strip()
    return name or os.path.splitext(filename)[0]


def split_upload_workbook(file_path, max_rows=200):
    """Split one converted workbook into government-sized upload batches."""
    workbook = openpyxl.load_workbook(file_path)
    try:
        worksheet = workbook['sheet1']
        data_rows = [
            row_number for row_number in range(3, worksheet.max_row + 1)
            if worksheet.cell(row=row_number, column=1).value
            and worksheet.cell(row=row_number, column=5).value
        ]
    finally:
        workbook.close()

    if len(data_rows) <= max_rows:
        return [file_path]

    batch_dir = os.path.join(os.path.dirname(file_path), '.upload_batches')
    os.makedirs(batch_dir, exist_ok=True)
    filename = os.path.basename(file_path)
    suffix = '_姜太公导出.xlsx'
    stem = filename[:-len(suffix)] if filename.endswith(suffix) else os.path.splitext(filename)[0]
    batch_paths = []
    for batch_number, offset in enumerate(range(0, len(data_rows), max_rows), start=1):
        selected_rows = data_rows[offset:offset + max_rows]
        batch_path = os.path.join(batch_dir, f'{stem}_第{batch_number:03d}批_姜太公导出.xlsx')
        shutil.copy2(file_path, batch_path)
        batch_workbook = openpyxl.load_workbook(batch_path)
        try:
            batch_sheet = batch_workbook['sheet1']
            last_row = selected_rows[-1]
            if last_row < batch_sheet.max_row:
                batch_sheet.delete_rows(last_row + 1, batch_sheet.max_row - last_row)
            first_row = selected_rows[0]
            if first_row > 3:
                batch_sheet.delete_rows(3, first_row - 3)
            batch_workbook.save(batch_path)
        finally:
            batch_workbook.close()
        batch_paths.append(batch_path)

    logger.info(
        "上传文件共 %s 条，已拆分为 %s 批，每批最多 %s 条",
        len(data_rows), len(batch_paths), max_rows,
    )
    return batch_paths


def count_upload_rows(file_path):
    workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    try:
        worksheet = workbook['sheet1']
        return sum(
            1 for row in worksheet.iter_rows(min_row=3, values_only=True)
            if row[0] and len(row) > 4 and row[4]
        )
    finally:
        workbook.close()


class Uploader:
    def __init__(self, config, driver=None):
        self.config = config
        self.browser_config = config['browser']
        self.driver = driver
        self.own_driver = driver is None
        self.long_wait = None
        self.short_wait = None
        self.target = None
        self.error_files = []
        self.results = []
        self.validate_only = bool(config.get('runtime', {}).get('validate_only', False))
        self.date_str = datetime.now().strftime('%Y%m%d')
        self.script_dir = os.path.dirname(os.path.abspath(__file__))

        self.output_dir = os.path.join(
            os.path.abspath(config['browser']['output_dir']),
            self.date_str
        )
        self.error_dir = os.path.join(
            os.path.abspath(config['browser'].get('error_dir', './数据转换/error')),
            self.date_str
        )
        os.makedirs(self.error_dir, exist_ok=True)

        # Keep the setting configurable, but bound a stalled import at ten
        # minutes by default instead of allowing an unbounded wait.
        timeout_minutes = config['browser'].get('upload_timeout', 10)
        # Safety cap: an old persisted setting of 15 minutes must not revive
        # the previous long wait while the task is stalled.
        self.upload_timeout = min(max(int(timeout_minutes), 1), 10) * 60
        configured_retries = config.get('runtime', {}).get('max_execution_retries', 3)
        try:
            configured_retries = int(configured_retries)
        except (TypeError, ValueError):
            configured_retries = 3
        self.max_execution_retries = min(max(configured_retries, 0), 3)

    def _init_driver(self):
        if self.driver is not None:
            return

        tmp_dir = os.path.join(self.error_dir, 'tmp')
        os.makedirs(tmp_dir, exist_ok=True)
        self.driver = create_driver(
            self.error_dir,
            headless=self.config.get('browser', {}).get('headless', True),
            profile_dir=browser_profile_dir("target", self.config['credentials']['target_site']['username']),
        )
        self.driver.set_page_load_timeout(PAGE_LOAD_TIMEOUT)
        self.driver.implicitly_wait(0)
        self.long_wait = WebDriverWait(self.driver, 15)
        self.short_wait = WebDriverWait(self.driver, 5)
        logger.info("Chromium 浏览器已启动")

    def _login(self, force_reauth=False):
        from target_login import TargetLogin
        self.target = TargetLogin(self.config)
        self.target.driver = self.driver
        self.target.long_wait = self.long_wait
        self.target.short_wait = self.short_wait

        try:
            self.target._init_driver = lambda: None
            self.target._init_driver()
            self.driver = self.target.driver
            self.long_wait = WebDriverWait(self.driver, 15)
            self.short_wait = WebDriverWait(self.driver, 5)

            if not force_reauth and self.target._reuse_existing_session():
                return
            if force_reauth:
                logger.warning("目标站跨域登录状态已过期，清理旧会话后重新登录")
                self.target.clear_persisted_session()
            if not self.target._start_feishu_listener():
                raise RuntimeError("飞书监听启动失败")
            if not self.target.login(reuse_session=False):
                raise RuntimeError("登录失败")
        except Exception:
            self.target._stop_feishu_listener()
            raise

    def run(self):
        try:
            self._init_driver()
            self._login()
            self._do_uploads()
            self._send_error_files_email()
            return self.results
        finally:
            self.close()

    def upload_only(self):
        self._init_driver()
        self._do_uploads()
        self._send_error_files_email()
        return self.results

    def _do_uploads(self):
        if not os.path.exists(self.output_dir):
            logger.error(f"输出目录不存在: {self.output_dir}")
            return self.results

        source_files = sorted([
            os.path.join(self.output_dir, filename) for filename in os.listdir(self.output_dir)
            if filename.endswith('.xlsx') and not filename.startswith('~$')
        ])
        xlsx_files = [
            batch_path for source_path in source_files
            for batch_path in split_upload_workbook(source_path)
        ]

        if not xlsx_files:
            logger.warning("没有需要上传的文件")
            return self.results

        logger.info(f"共 {len(xlsx_files)} 个文件待上传:")
        for file_path in xlsx_files:
            logger.info(f"  - {os.path.basename(file_path)}")

        session_needs_restart = False
        force_reauth = False
        for idx, file_path in enumerate(xlsx_files):
            filename = os.path.basename(file_path)
            project_name = extract_project_name(file_path)
            logger.info(f"[{idx + 1}/{len(xlsx_files)}] 开始上传: {project_name}")

            for attempt in range(self.max_execution_retries + 1):
                try:
                    if session_needs_restart:
                        if force_reauth:
                            self._restart_session(force_reauth=True)
                        else:
                            self._restart_session()
                        session_needs_restart = False
                        force_reauth = False
                    result = self._upload_single_file(
                        file_path,
                        project_name,
                        is_first=True,
                    )
                    self.results.append(result)
                    break
                except Exception as e:
                    self._log_page_state("上传失败现场")
                    session_needs_restart = True
                    force_reauth = isinstance(e, TargetSessionExpired)
                    error = str(e)
                    is_business_rejection = error.startswith(
                        "政府平台未接受该人员"
                    )
                    if attempt < self.max_execution_retries and not is_business_rejection:
                        logger.warning(
                            "上传执行异常 %s，准备自动重试 %s/%s: %s",
                            project_name,
                            attempt + 1,
                            self.max_execution_retries,
                            e,
                            exc_info=True,
                        )
                        continue
                    if is_business_rejection:
                        logger.info("政府平台已返回业务拒绝明细，不再重试上传")
                    logger.error(f"上传失败 {project_name}: {e}", exc_info=True)
                    batch_rows = count_upload_rows(file_path)
                    self.results.append({
                        'project_name': project_name,
                        'status': 'failed',
                        'error': error,
                        # 对方平台明确拒绝人员是业务结果；找不到按钮、
                        # 页面未跳转等其他异常是自动化执行错误，必须让任务失败。
                        'execution_error': not is_business_rejection,
                        'total_rows': batch_rows,
                        'success_rows': 0,
                        'failure_rows': batch_rows,
                    })
                    break

        logger.info(f"全部上传完成，失败/错误文件: {len(self.error_files)}")
        return self.results

    def _restart_session(self, force_reauth=False):
        if not self.own_driver:
            raise RuntimeError("外部浏览器会话无法自动重启")
        self.close()
        self.driver = None
        self.long_wait = None
        self.short_wait = None
        self.target = None
        self._init_driver()
        self._login(force_reauth=force_reauth)

    def _upload_single_file(self, file_path, project_name, is_first=False):
        if is_first:
            self._navigate_to_service()

        self._handle_site_selection()

        self._click_online_handle()

        self._handle_authorization_confirmation()

        self._handle_draft_popup()

        if self._is_upload_form_visible():
            logger.info("已进入在线填表页，跳过备案类型选择和确认")
        else:
            self._select_registration_type()
            self._confirm_selection()

        self._wait_for_upload_form()
        self._wait_for_page_ready()

        self._click_batch_import()

        self._upload_file(file_path)

        self._wait_for_upload_complete()

        form_handle = self.driver.current_window_handle

        upload_result = self._read_upload_results(project_name)

        handles = self.driver.window_handles
        if len(handles) > 1:
            logger.info("关闭错误下载弹出的额外标签页...")
            for h in handles:
                if h != form_handle:
                    self.driver.switch_to.window(h)
                    try:
                        self.driver.close()
                    except Exception:
                        pass
            self.driver.switch_to.window(form_handle)

        if upload_result.get('already_exists'):
            logger.info("该人员在政府平台已存在，按测试成功处理，不再重复提交")
            self._click_home()
            return {**upload_result, 'project_name': project_name, 'status': 'success'}

        if upload_result.get('total_rows', 0) and upload_result.get('success_rows') == 0:
            raise RuntimeError("政府平台未接受该人员，且未返回‘已存在’结果")

        if self.validate_only:
            logger.info("上传校验测试完成，按配置停在最终提交前")
            # Return to the service home so a validation run containing more
            # than one project can start the next upload from a known page.
            self._click_home()
            return {**upload_result, 'project_name': project_name, 'status': 'validated'}

        self._click_next_step()

        self._wait_for_page_ready()

        self._click_submit()

        if not self._check_submit_success():
            raise RuntimeError("未检测到最终提交成功提示")

        if self.error_files:
            logger.info(f"当前累计错误文件: {len(self.error_files)}")

        self._click_home()
        return {**upload_result, 'project_name': project_name, 'status': 'success'}

    def _navigate_to_service(self):
        logger.info("点击'建筑项目人员工伤参保登记'")
        for _ in range(2):
            try:
                self.driver.get(HOME_URL)
                break
            except TimeoutException:
                logger.warning("页面加载超时，尝试停止加载并重试...")
                try:
                    self.driver.execute_script('window.stop();')
                except Exception:
                    pass
                time.sleep(2)
        try:
            self.long_wait.until(
                lambda d: d.execute_script('return document.readyState') in ('complete', 'interactive')
            )
        except TimeoutException:
            logger.warning("页面加载超时，继续尝试...")
        time.sleep(2)

        selectors = [
            (By.XPATH, "//span[contains(text(),'建筑项目人员工伤参保登记')]"),
            (By.XPATH, "//*[contains(text(),'建筑项目') and contains(text(),'工伤')]"),
            (By.XPATH, "//p[contains(text(),'建筑项目人员')]"),
            (By.XPATH, "//*[contains(text(),'工伤参保登记')]"),
        ]
        el = _find_first(self.driver, selectors)
        if el:
            original_handles = set(self.driver.window_handles)
            original_url = self.driver.current_url
            self._click(el)
            logger.info("已点击'建筑项目人员工伤参保登记'")
            self._wait_for_search_result_navigation(original_handles, original_url)
            return

        logger.info("首页未找到直接链接，通过搜索框搜索...")
        search_input = _find_first(self.driver, [
            (By.XPATH, "//form//input[@type='text']"),
            (By.XPATH, "//input[contains(@placeholder,'搜索')]"),
            (By.XPATH, "//input[@type='text' or @type='search']"),
            (By.XPATH, "//form//input"),
        ])
        if not search_input:
            raise RuntimeError("未找到搜索框")

        search_input.clear()
        search_input.send_keys('建筑项目人员工伤参保登记')
        time.sleep(0.5)

        search_btn = _find_first(self.driver, [
            (By.XPATH, "//input[@type='submit']"),
            (By.XPATH, "//input[@value='搜索']"),
            (By.XPATH, "//input[contains(@value,'搜索')]"),
            (By.XPATH, "//form//input[@type='submit']"),
            (By.XPATH, "//form//*[@type='submit']"),
            (By.XPATH, "//button[@type='submit']"),
            (By.XPATH, "//span[contains(text(),'搜索')]/parent::*"),
        ])
        if search_btn:
            self._click(search_btn)
            logger.info("已点击搜索按钮")
        else:
            logger.info("未找到搜索按钮，尝试 Enter 提交")
            search_input.send_keys('\n')

        logger.info("等待搜索页跳转...")
        for _ in range(15):
            time.sleep(1)
            if len(self.driver.window_handles) > 1:
                original = self.driver.current_window_handle
                for handle in self.driver.window_handles:
                    if handle != original:
                        self.driver.switch_to.window(handle)
                        logger.info(f"切换到新窗口: {self.driver.current_url}")
                        break
                break
            if 'search.zj.gov.cn' in self.driver.current_url:
                logger.info(f"搜索页 URL 已变化: {self.driver.current_url}")
                break

        logger.info(f"搜索后 URL: {self.driver.current_url}")
        logger.info(f"搜索后标题: {self.driver.title}")

        try:
            self.long_wait.until(
                lambda d: d.execute_script('return document.readyState') in ('complete', 'interactive')
            )
        except TimeoutException:
            logger.warning("搜索页加载超时")
        time.sleep(2)

        self._switch_to_main_iframe()

        logger.info("查找搜索结果 (使用绝对 XPath)...")
        xpaths = [
            "/html/body/div[1]/form//a[contains(.,'建筑项目人员工伤参保登记')]",
            "/html/body/div[1]/form/div[7]/div/div[1]/div/div[1]/div/a",
            "/html/body/div[1]//a[contains(.,'建筑项目')]",
            "//a[contains(.,'建筑项目人员工伤参保登记')]",
            "//a[contains(.,'建筑项目')]",
            "//*[contains(text(),'建筑项目人员工伤参保登记')]/ancestor::a",
        ]

        for xpath in xpaths:
            try:
                el = self.long_wait.until(EC.element_to_be_clickable((By.XPATH, xpath)))
                original_handles = set(self.driver.window_handles)
                original_url = self.driver.current_url
                self._click(el)
                logger.info(f"已点击搜索结果 ({xpath})")
                self._wait_for_search_result_navigation(original_handles, original_url, strict=True)
                return
            except TimeoutException:
                continue

        logger.info("JS 查找兜底...")
        original_handles = set(self.driver.window_handles)
        original_url = self.driver.current_url
        found = self._find_and_click_result_js('建筑项目人员工伤参保登记')
        if found:
            self._wait_for_search_result_navigation(original_handles, original_url, strict=True)
            return

        raise RuntimeError("未找到搜索结果")

    def _wait_for_search_result_navigation(self, original_handles, original_url, strict=False):
        logger.info("等待办事页面打开...")
        for _ in range(10):
            time.sleep(1)
            current_handles = set(self.driver.window_handles)
            new_handles = current_handles - set(original_handles)
            if new_handles:
                handle = new_handles.pop()
                self.driver.switch_to.window(handle)
                logger.info(f"切换到新页面: {self.driver.current_url}")
                try:
                    self.long_wait.until(
                        lambda d: d.execute_script('return document.readyState') in ('complete', 'interactive')
                    )
                except TimeoutException:
                    pass
                time.sleep(1)
                return True
            if self.driver.current_url != original_url:
                logger.info(f"当前窗口已跳转到办事页面: {self.driver.current_url}")
                time.sleep(1)
                return True
            if _find_first(self.driver, [
                (By.XPATH, "//*[normalize-space(.)='同意授权']"),
                (By.XPATH, "//button[contains(.,'在线办理')]"),
                (By.XPATH, "//a[contains(.,'在线办理')]"),
                (By.XPATH, "//li[contains(text(),'宁波')]"),
            ]):
                logger.info("当前窗口已渲染办事或授权页面")
                return True

        message = f"点击办事入口后页面未跳转，仍停留在: {self.driver.current_url}"
        logger.error(message)
        if strict:
            raise RuntimeError(message)
        return False

    def _handle_site_selection(self):
        logger.info("检查站点选择弹窗")
        time.sleep(1)

        els = self.driver.find_elements(By.XPATH, "//li[contains(text(),'宁波')]")
        visible = [e for e in els if e.is_displayed()]
        if not visible:
            logger.info("无站点选择弹窗")
            return

        logger.info("检测到站点选择弹窗，选择'宁波市'")
        self._click(visible[0])
        time.sleep(0.5)

        btn = _find_first(self.driver, [
            (By.XPATH, "//button[text()='确定']"),
            (By.XPATH, "//button[contains(text(),'确定')]"),
        ])
        if btn:
            original_handles = set(self.driver.window_handles)
            self._click(btn)
            logger.info("已点击'确定'，等待新窗口...")
            for _ in range(8):
                time.sleep(0.5)
                new_handles = set(self.driver.window_handles) - original_handles
                if new_handles:
                    self.driver.switch_to.window(new_handles.pop())
                    logger.info(f"已切换到新窗口: {self.driver.current_url}")
                    self.long_wait.until(
                        lambda d: d.execute_script('return document.readyState') in ('complete', 'interactive')
                    )
                    time.sleep(1)
                    return
            logger.info("确定点击后无新窗口")

    def _click_online_handle(self):
        logger.info("点击'在线办理'")
        selectors = [
            (By.CSS_SELECTOR, ".title-but > button.font-size-16"),
            (By.CSS_SELECTOR, ".bl-but.font-size-20"),
            (By.CSS_SELECTOR, ".title-but"),
            (By.XPATH, "//button[contains(text(),'在线办理')]"),
            (By.XPATH, "//span[contains(text(),'在线办理')]/parent::button"),
            (By.XPATH, "//a[contains(.,'在线办理')]"),
            (By.XPATH, "//div[normalize-space(text())='在线办理']"),
            (By.XPATH, "//button[contains(@class,'handle')]"),
        ]
        el = None
        deadline = time.time() + 60
        last_log = 0
        while time.time() < deadline:
            if self._target_session_expired():
                raise TargetSessionExpired(
                    "目标政务网跨域登录状态已过期（用户中心要求重新登录）"
                )
            el = _find_first(self.driver, selectors)
            if el:
                break
            elapsed = int(60 - (deadline - time.time()))
            if elapsed >= last_log + 10:
                logger.info(f"等待办事指南加载在线办理按钮... ({elapsed}秒)")
                last_log = elapsed
            time.sleep(0.5)
        if not el:
            raise RuntimeError("办事指南等待 60 秒仍未找到'在线办理'")

        original_handles = set(self.driver.window_handles)
        original_url = self.driver.current_url
        self._click(el)
        logger.info("已点击'在线办理'，等待办理流程打开...")
        deadline = time.time() + 45
        last_log = 0
        while time.time() < deadline:
            time.sleep(0.5)
            new_handles = set(self.driver.window_handles) - original_handles
            if new_handles:
                self.driver.switch_to.window(new_handles.pop())
                logger.info(f"已切换到在线办理窗口: {self.driver.current_url}")
                try:
                    self.long_wait.until(
                        lambda d: d.execute_script('return document.readyState') in ('complete', 'interactive')
                    )
                except TimeoutException:
                    logger.warning("在线办理窗口加载超时，继续识别页面状态")
                return

            if self._is_upload_form_visible():
                logger.info(f"当前窗口已直接进入在线填表页: {self.driver.current_url}")
                return
            # 办事指南本身也可能渲染备案类型文案或残留弹窗元素。
            # 只有当前窗口 URL 已发生跳转时，才能以这些元素判定已进入办理流程。
            if self.driver.current_url != original_url and _find_first(self.driver, [
                (By.XPATH, "//*[normalize-space(.)='同意授权']"),
                (By.XPATH, "//*[normalize-space(.)='建筑项目人员备案登记']"),
                (By.XPATH, "//*[contains(text(),'草稿') and (self::div or self::span)]"),
            ]):
                logger.info(f"当前窗口已进入办理流程: {self.driver.current_url}")
                return

            elapsed = int(45 - (deadline - time.time()))
            if elapsed >= last_log + 10:
                logger.info(
                    "等待在线办理页面跳转... (%s秒，当前 URL: %s)",
                    elapsed,
                    self.driver.current_url,
                )
                last_log = elapsed

        if self.driver.current_url == original_url:
            raise RuntimeError(
                f"点击在线办理后 45 秒仍停留在办事指南页: {self.driver.current_url}"
            )
        raise RuntimeError(
            f"点击在线办理后未识别到授权页、备案类型选择或在线填表页: {self.driver.current_url}"
        )

    def _target_session_expired(self):
        current_url = (self.driver.current_url or '').lower()
        if '#/mymessage' not in current_url and 'zjucenter' not in current_url:
            return False
        return _find_first(self.driver, [
            (By.XPATH, "//*[self::a or self::button or self::span][normalize-space(.)='立即登录']"),
            (By.XPATH, "//*[self::a or self::button][normalize-space(.)='登录']"),
        ]) is not None

    def _log_page_state(self, label):
        try:
            current_url = self.driver.current_url
            title = self.driver.title
            evidence_root = self.browser_config.get('diagnostics_dir') or os.path.join(
                os.environ.get('REPORT_FORWARD_LOCAL_STORAGE', '/data/artifacts'),
                'diagnostics', self.date_str,
            )
            os.makedirs(evidence_root, exist_ok=True)
            safe_label = re.sub(r'[^0-9A-Za-z一-龥_-]+', '_', label).strip('_') or 'page'
            stamp = datetime.now().strftime('%H%M%S_%f')
            prefix = os.path.join(evidence_root, f'{stamp}_{safe_label}')
            screenshot_path = f'{prefix}.png'
            html_path = f'{prefix}.html'
            self.driver.save_screenshot(screenshot_path)
            with open(html_path, 'w', encoding='utf-8') as html_file:
                html_file.write(self.driver.page_source or '')
            actions = self.driver.execute_script("""
                return Array.from(document.querySelectorAll('button,a,[role="button"],input[type="button"],input[type="submit"]'))
                    .filter(function (el) { return !!(el.offsetWidth || el.offsetHeight || el.getClientRects().length); })
                    .map(function (el) { return (el.innerText || el.value || el.textContent || '').trim(); })
                    .filter(Boolean)
                    .slice(0, 40);
            """)
            logger.info(
                f"{label} — URL: {current_url}, 标题: {title}, 可见操作: {actions}; "
                f"现场证据已保存: {screenshot_path}, {html_path}"
            )
        except Exception as error:
            logger.warning(f"记录页面现场失败: {error}")

    def _handle_authorization_confirmation(self):
        """Accept the optional government authorization page before entering the form."""
        logger.info("检查是否出现授权确认页")
        agree_selectors = [
            (By.XPATH, "//button[normalize-space(.)='同意授权']"),
            (By.XPATH, "//a[normalize-space(.)='同意授权']"),
            (By.XPATH, "//input[@type='button' and @value='同意授权']"),
            (By.XPATH, "//*[contains(@class,'button') and normalize-space(.)='同意授权']"),
        ]
        agree = None
        for _ in range(6):
            agree = _find_first(self.driver, agree_selectors)
            if agree:
                break
            time.sleep(0.5)
        if not agree:
            logger.info("未出现授权确认页，继续办理")
            return False

        logger.info("检测到授权确认页，勾选以后无需授权并同意授权")
        checkbox = _find_first(self.driver, [
            (By.XPATH, "//input[@type='checkbox']"),
            (By.XPATH, "//*[@role='checkbox']"),
            (By.XPATH, "//*[contains(normalize-space(.),'以后无需授权')]/preceding::input[@type='checkbox'][1]"),
            (By.XPATH, "//*[contains(normalize-space(.),'以后无需授权') and (self::label or self::span)]"),
        ])
        if checkbox:
            try:
                if not getattr(checkbox, 'is_selected', lambda: False)():
                    self._click(checkbox)
                    logger.info("已勾选'以后无需授权可直接使用您的信息'")
            except Exception as error:
                logger.warning(f"勾选授权选项失败，仍尝试同意授权: {error}")
        else:
            logger.warning("未找到'以后无需授权'勾选框，仍尝试同意授权")

        self._click(agree)
        logger.info("已点击'同意授权'")
        time.sleep(2)
        return True

    def _handle_draft_popup(self):
        logger.info("检查草稿弹窗")

        for _ in range(10):
            els = self.driver.find_elements(By.XPATH, "//*[contains(text(),'草稿')]")
            if any(e.is_displayed() for e in els):
                break
            time.sleep(1)
        else:
            logger.info("无草稿弹窗")
            return

        logger.info("检测到草稿弹窗，点击'不使用'")
        selectors = [
            (By.XPATH, "//span[text()='不使用']"),
            (By.XPATH, "//span[contains(text(),'不使用')]/parent::button"),
            (By.XPATH, "//button[contains(.,'不使用')]"),
        ]
        btn = _find_first(self.driver, selectors)
        if btn:
            self._click(btn)
            logger.info("已点击'不使用'")
            time.sleep(2)
        else:
            logger.warning("未找到'不使用'按钮")

    def _is_upload_form_visible(self):
        if _find_first(self.driver, [
            (By.XPATH, "//input[@type='file']"),
            (By.XPATH, "//button[.//*[normalize-space(.)='批量导入'] or normalize-space(.)='批量导入']"),
            (By.XPATH, "//*[normalize-space(.)='建筑项目人员备案信息']"),
        ]):
            return True
        unit_info = _find_first(self.driver, [
            (By.XPATH, "//*[normalize-space(.)='单位信息']"),
        ])
        online_form = _find_first(self.driver, [
            (By.XPATH, "//*[normalize-space(.)='在线填表']"),
            (By.XPATH, "//*[contains(normalize-space(.),'备案开始时间')]"),
        ])
        return bool(unit_info and online_form)

    def _registration_confirm_button(self):
        return _find_first(self.driver, [
            (By.XPATH, "//*[@role='dialog']//button[translate(normalize-space(.), ' ', '')='确认' or translate(normalize-space(.), ' ', '')='确定']"),
            (By.XPATH, "//*[contains(@class,'dialog') or contains(@class,'modal')]//button[translate(normalize-space(.), ' ', '')='确认' or translate(normalize-space(.), ' ', '')='确定']"),
            # 政务网新版“情形选择”是独立页而非弹窗，主按钮 DOM
            # 文本为“确  定”。限定在建筑项目人员备案选项后查找。
            (By.XPATH, "//*[normalize-space(.)='建筑项目人员备案登记']/following::button[contains(@class,'next-btn-primary')][1]"),
            (By.XPATH, "//button[translate(normalize-space(.), ' ', '')='确认' or translate(normalize-space(.), ' ', '')='确定']"),
        ])

    def _select_registration_type(self):
        if self._is_upload_form_visible():
            logger.info("已进入在线填表页，无需选择备案类型")
            return False
        logger.info("选择'建筑项目人员备案登记'")
        # 先确认页面上确实存在情形选择页/弹窗的确定按钮，避免把办事指南标题
        # “建筑项目人员工伤参保登记”误当成备案类型选项。
        if not self._registration_confirm_button():
            raise RuntimeError(
                f"未出现备案类型选择页或弹窗，当前页面: {self.driver.current_url}"
            )
        selectors = [
            (By.XPATH, "//*[@role='dialog']//*[normalize-space(.)='建筑项目人员备案登记']"),
            (By.XPATH, "//*[contains(@class,'dialog') or contains(@class,'modal')]//*[normalize-space(.)='建筑项目人员备案登记']"),
            (By.XPATH, "//span[normalize-space(.)='建筑项目人员备案登记']"),
            (By.XPATH, "//label[normalize-space(.)='建筑项目人员备案登记']"),
        ]
        el = _find_first(self.driver, selectors)
        if not el:
            raise RuntimeError("备案类型选择页或弹窗中未找到'建筑项目人员备案登记'")

        self._click(el)
        logger.info("已选择'建筑项目人员备案登记'")
        time.sleep(1)
        return True

    def _confirm_selection(self):
        if self._is_upload_form_visible():
            logger.info("备案类型选择后已直接进入在线填表页，跳过确认")
            return False
        logger.info("点击备案类型选择弹窗的确认按钮")
        el = self._registration_confirm_button()
        if not el:
            raise RuntimeError(
                f"备案类型选择页或弹窗仍在，但未找到确定按钮: {self.driver.current_url}"
            )

        self._click(el)
        logger.info("已点击确认")
        return True

    def _wait_for_upload_form(self, timeout=30):
        deadline = time.time() + timeout
        while time.time() < deadline:
            if self._is_upload_form_visible():
                logger.info("在线填表页已就绪")
                return
            time.sleep(0.5)
        raise RuntimeError(
            f"办理流程未进入在线填表页，当前页面: {self.driver.current_url}"
        )

    def _ensure_zjzwfw_window(self):
        handles = self.driver.window_handles
        logger.info(f"当前窗口数: {len(handles)}, 当前 URL: {self.driver.current_url}")

        if 'zjzwfw.gov.cn' in self.driver.current_url:
            return

        for handle in handles:
            try:
                self.driver.switch_to.window(handle)
                if 'zjzwfw.gov.cn' in self.driver.current_url:
                    logger.info(f"切换到 zjzwfw 窗口: {self.driver.current_url}")
                    return
            except Exception:
                continue

        logger.warning(f"未找到 zjzwfw 窗口，保持当前: {self.driver.current_url}")

    def _switch_to_main_iframe(self):
        try:
            self.driver.switch_to.default_content()
        except Exception:
            pass

    def _find_and_click_result_js(self, text):
        try:
            self.driver.switch_to.default_content()
        except Exception:
            pass

        iframes = self.driver.find_elements(By.TAG_NAME, "iframe")
        for iframe in iframes:
            try:
                self.driver.switch_to.frame(iframe)
                found = self._js_find_and_click(text)
                if found:
                    logger.info(f"在 iframe 中找到并点击了: {text}")
                    return found
            except Exception:
                pass
            try:
                self.driver.switch_to.default_content()
            except Exception:
                pass

        try:
            self.driver.switch_to.default_content()
        except Exception:
            pass

        return self._js_find_and_click(text)

    def _js_find_and_click(self, text):
        js_code = """
        var links = document.querySelectorAll('a');
        for (var i = 0; i < links.length; i++) {
            var link = links[i];
            if (link.textContent.indexOf(arguments[0]) >= 0) {
                link.click();
                return true;
            }
        }
        var clickables = document.querySelectorAll('span, p, div, li');
        for (var j = 0; j < clickables.length; j++) {
            var el = clickables[j];
            if (el.textContent.indexOf(arguments[0]) >= 0) {
                var a = el.closest('a');
                if (a) {
                    a.click();
                    return true;
                }
                el.click();
                return true;
            }
        }
        return false;
        """
        try:
            found = self.driver.execute_script(js_code, text)
            if found:
                logger.info(f"JS 查找并点击成功: {text}")
            return found
        except Exception as e:
            logger.error(f"JS 查找异常: {e}")
            return False

    def _wait_for_page_ready(self):
        try:
            self.long_wait.until(
                lambda d: d.execute_script('return document.readyState') in ('complete', 'interactive')
            )
        except TimeoutException:
            pass
        time.sleep(1)

    def _click_batch_import(self):
        logger.info("准备批量导入...")

    def _close_file_dialog(self):
        try:
            subprocess.run([
                'powershell', '-Command',
                'Add-Type -AssemblyName System.Windows.Forms; [System.Windows.Forms.SendKeys]::SendWait("{ESC}")'
            ], timeout=5, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
            logger.info("已发送 ESC 尝试关闭文件对话框")
            time.sleep(1)
        except Exception as e:
            logger.warning(f"关闭对话框失败: {e}")

    def _upload_file(self, file_path):
        logger.info(f"上传文件: {file_path}")
        abs_path = os.path.abspath(file_path)

        file_input = _find_first(self.driver, [(By.XPATH, "//input[@type='file']")])
        if not file_input:
            logger.info("未找到文件input，点击'批量导入'触发...")
            el = _find_first(self.driver, [
                (By.XPATH, "//button[contains(.,'批量导入')]"),
                (By.XPATH, "//span[contains(text(),'批量导入')]/parent::button"),
                (By.XPATH, "//*[contains(text(),'批量导入')]"),
            ])
            if el:
                self._click(el)
                logger.info("已点击'批量导入'")
                time.sleep(2)
                self._close_file_dialog()

            file_input = _find_first(self.driver, [(By.XPATH, "//input[@type='file']")])

        if not file_input:
            try:
                file_input = self.long_wait.until(
                    EC.presence_of_element_located((By.XPATH, "//input[@type='file']"))
                )
            except TimeoutException:
                raise RuntimeError("未找到文件上传输入框")

        file_input.send_keys(abs_path)
        logger.info(f"文件已选择: {os.path.basename(abs_path)}")
        time.sleep(1)

    def _wait_for_upload_complete(self):
        logger.info("等待上传完成...")
        start = time.time()

        progress_selectors = [
            (By.XPATH, "//div[contains(@class,'progress')]"),
            (By.XPATH, "//*[contains(text(),'上传中')]"),
            (By.XPATH, "//*[contains(text(),'导入中')]"),
            (By.XPATH, "//*[contains(text(),'处理中')]"),
        ]

        result_selectors = [
            (By.XPATH, "//*[contains(text(),'共计')]"),
            (By.XPATH, "//*[contains(text(),'成功')]"),
            (By.XPATH, "//*[contains(text(),'上传') and contains(text(),'行')]"),
        ]

        while time.time() - start < self.upload_timeout:
            result_el = _find_first(self.driver, result_selectors)
            if result_el:
                logger.info("检测到上传结果")
                time.sleep(1)
                return

            time.sleep(1)

            elapsed = int(time.time() - start)
            if elapsed % 15 < 2:
                progress_el = _find_first(self.driver, progress_selectors, check_displayed=True)
                if progress_el:
                    logger.info(f"上传进行中... (已等待 {elapsed}秒)")
                else:
                    logger.info(f"等待上传响应... (已等待 {elapsed}秒)")

        logger.error(f"上传超时 ({self.upload_timeout}秒)，停止等待并让上层按失败重试策略处理")
        raise TimeoutError(f"目标站上传超时（{self.upload_timeout}秒）")

    def _read_upload_results(self, project_name):
        logger.info("读取上传结果")

        text = self.driver.find_element(By.TAG_NAME, "body").text

        total_match = re.search(r'共计[上传导入]*\s*(\d+)\s*行', text)
        success_match = re.search(r'成功\s*(\d+)\s*行', text)

        total_rows = total_match.group(1) if total_match else '?'
        success_rows = success_match.group(1) if success_match else '?'

        logger.info(f"上传结果 — 共计: {total_rows}行, 成功: {success_rows}行")

        error_links = self.driver.find_elements(By.XPATH, "//a[contains(@href,'.xlsx') or contains(@href,'.xls')]")
        download_links = [a for a in error_links if a.is_displayed()]

        if not download_links:
            error_links = self.driver.find_elements(
                By.XPATH, "//a[contains(text(),'下载') or contains(text(),'明细') or contains(text(),'错误')]"
            )
            download_links = [a for a in error_links if a.is_displayed()]

        downloaded_error_files = []
        if total_rows != success_rows and download_links:
            logger.info(f"发现 {len(download_links)} 个错误明细下载链接")
            downloaded_error_files = self._download_error_file(download_links[0], project_name)
        else:
            logger.info("未发现错误明细文件，或上传全部成功")

        result_text = text
        for error_file in downloaded_error_files:
            result_text += "\n" + self._read_workbook_text(error_file)
        normalized = normalize_upload_result(total_rows, success_rows, result_text)
        person_results = extract_error_person_results(downloaded_error_files)
        if person_results:
            normalized['person_results'] = person_results
            normalized['person_details_available'] = (
                normalized['failure_rows'] is not None
                and len(person_results) == normalized['failure_rows']
            )
        if normalized['already_exists']:
            logger.info("政府平台返回该人员已存在，计为测试成功")
        elif normalized['person_details_available']:
            logger.info(
                "已从政府错误明细对应到 %s 条失败人员，剩余 %s 条为成功",
                len(person_results), normalized['success_rows'],
            )
        elif normalized['total_rows'] and normalized['total_rows'] > 1 and normalized['failure_rows']:
            logger.warning(
                "政府平台批量结果为成功 %s 条、失败 %s 条；未返回可可靠对应到个人的成功名单",
                normalized['success_rows'], normalized['failure_rows'],
            )
        return normalized

    def _download_error_file(self, link_element, project_name):
        files_before = set(os.listdir(self.error_dir))
        timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
        downloaded = []

        try:
            self._click(link_element)
            logger.info("已点击错误明细下载链接")
        except Exception as e:
            logger.error(f"点击下载链接失败: {e}")
            return downloaded

        time.sleep(2)
        for _ in range(30):
            time.sleep(1)
            current_files = set(os.listdir(self.error_dir))
            new_files = current_files - files_before
            completed = [f for f in new_files if f.endswith('.xlsx') and not f.startswith('~$')]
            downloading = [f for f in new_files if f.endswith('.crdownload') or f.endswith('.tmp')]

            if completed:
                new_name = f"{timestamp}_{project_name}_上传错误明细.xlsx"
                new_path = os.path.join(self.error_dir, new_name)
                if os.path.exists(new_path):
                    os.remove(new_path)
                shutil.move(os.path.join(self.error_dir, completed[0]), new_path)
                self.error_files.append(new_path)
                downloaded.append(new_path)
                logger.info(f"错误明细已下载: {new_name}")
                break

            if not downloading:
                time.sleep(0.5)

        current_files = set(os.listdir(self.error_dir))
        remaining = current_files - files_before
        for f in remaining:
            if f.endswith('.xlsx') and not f.startswith('~$'):
                name_without_ext = os.path.splitext(f)[0]
                if any(kw in name_without_ext for kw in ('上传错误', '错误明细', project_name[:6])):
                    continue
                new_name = f"{timestamp}_{project_name}_上传错误明细_{f}"
                new_path = os.path.join(self.error_dir, new_name)
                try:
                    shutil.move(os.path.join(self.error_dir, f), new_path)
                    self.error_files.append(new_path)
                    downloaded.append(new_path)
                    logger.info(f"额外错误文件已重命名: {new_name}")
                except Exception:
                    pass
        return downloaded

    def _read_workbook_text(self, path):
        try:
            workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
            try:
                values = []
                for sheet in workbook.worksheets:
                    for row in sheet.iter_rows(max_row=200, max_col=50, values_only=True):
                        values.extend(str(value) for value in row if value is not None)
                return "\n".join(values)
            finally:
                workbook.close()
        except Exception as error:
            logger.warning(f"读取错误明细失败: {error}")
            return ""

    def _click_next_step(self):
        logger.info("点击'下一步'")
        selectors = [
            (By.XPATH, "//button[contains(text(),'下一步')]"),
            (By.XPATH, "//span[text()='下一步']/parent::button"),
            (By.XPATH, "//span[contains(text(),'下一步')]"),
        ]
        el = _find_first(self.driver, selectors)
        if not el:
            try:
                el = self.long_wait.until(
                    EC.element_to_be_clickable((By.XPATH, "//button[contains(.,'下一步')]"))
                )
            except TimeoutException:
                raise RuntimeError("未找到'下一步'")

        self._click(el)
        logger.info("已点击'下一步'，等待跳转...")
        time.sleep(2)
        self.long_wait.until(
            lambda d: d.execute_script('return document.readyState') in ('complete', 'interactive')
        )
        time.sleep(1)
        logger.info(f"当前页面标题: {self.driver.title}")

    def _click_submit(self):
        logger.info("等待'信息确认'页面...")
        for _ in range(15):
            if '信息确认' in self.driver.title or '确认' in self.driver.title:
                break
            time.sleep(1)

        logger.info(f"当前页面: {self.driver.title}")
        time.sleep(2)

        logger.info("点击'提交'")
        selectors = [
            (By.XPATH, "/html/body/div[4]/div/div/div/div[3]/div/div[2]/button[2]"),
            (By.XPATH, "//button[.//span[contains(text(),'提交')]]"),
            (By.XPATH, "//button[contains(text(),'提交')]"),
            (By.XPATH, "//span[text()='提交']/parent::button"),
        ]
        for by, value in selectors:
            try:
                el = self.long_wait.until(EC.element_to_be_clickable((by, value)))
                self._click(el)
                logger.info("已点击提交")
                time.sleep(3)
                return
            except TimeoutException:
                continue

        raise RuntimeError("未找到'提交'")

    def _check_submit_success(self):
        logger.info("检查提交结果")
        start = time.time()
        while time.time() - start < 15:
            els = self.driver.find_elements(By.XPATH, "//*[contains(text(),'申请提交成功')]")
            visible = [e for e in els if e.is_displayed()]
            if visible:
                logger.info("申请提交成功!")
                return True

            els = self.driver.find_elements(By.XPATH, "//*[contains(text(),'提交成功')]")
            visible = [e for e in els if e.is_displayed()]
            if visible:
                logger.info("提交成功!")
                return True

            time.sleep(1)

        logger.warning("未检测到提交成功提示")
        return False

    def _click_home(self):
        logger.info("点击首页")
        selectors = [
            (By.XPATH, "//a[contains(text(),'首页')]"),
            (By.XPATH, "//span[contains(text(),'首页')]"),
            (By.XPATH, "//*[contains(text(),'首页')]"),
        ]
        el = _find_first(self.driver, selectors)
        if el:
            self._click(el)
            logger.info("已点击首页")
            time.sleep(3)
        else:
            logger.info("未找到首页链接，直接导航")
            for _ in range(2):
                try:
                    self.driver.get(HOME_URL)
                    break
                except TimeoutException:
                    logger.warning("页面加载超时，尝试停止加载并重试...")
                    try:
                        self.driver.execute_script('window.stop();')
                    except Exception:
                        pass
                    time.sleep(2)
            time.sleep(3)

    def _send_error_files_email(self):
        if not self.error_files:
            logger.info("没有错误文件需要发送")
            return

        if not self.config.get('email', {}).get('enabled', False):
            logger.info("邮件通知未启用，错误文件已由报送中心留存")
            return

        from email_sender import send_email

        subject = f"上传错误明细 - {self.date_str}"
        body_parts = [f"日期: {self.date_str}", f"共计 {len(self.error_files)} 个错误明细文件:", ""]
        for f in self.error_files:
            body_parts.append(f"  - {os.path.basename(f)}")

        send_email(self.config, subject, '\n'.join(body_parts), self.error_files)

    def _click(self, element):
        try:
            element.click()
        except ElementClickInterceptedException:
            try:
                self.driver.execute_script("arguments[0].click();", element)
            except Exception:
                pass
        except Exception:
            self.driver.execute_script("arguments[0].click();", element)

    def close(self):
        if self.target:
            self.target._stop_feishu_listener()
        if self.driver and self.own_driver:
            try:
                close_driver(self.driver)
                logger.info("浏览器已关闭")
            except Exception as e:
                logger.warning(f"关闭浏览器异常: {e}")
